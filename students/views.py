from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Avg, Count, Max, Q
from .models import Student, Result, Subject
from .forms import StudentForm, ResultForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
import pandas as pd
from openpyxl import Workbook
from django.core.paginator import Paginator  

@login_required
def student_list(request):
    # QUERYSET examples
    students = Student.objects.all().order_by('-created_at')
    
    # Search functionality using Q objects
    search_query = request.GET.get('search')
    if search_query:
        students = students.filter(
            Q(name__icontains=search_query) |
            Q(roll_no__icontains=search_query) |
            Q(class_name__icontains=search_query)
        )
    
    # PAGINATION - Har page pe 10 students
    paginator = Paginator(students, 10)  # 10 students per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    total_students = Student.objects.count()
    
    context = {
        'students': page_obj,  # ← Ye page_obj hai (students nahi)
        'total_students': total_students,
        'search_query': search_query,
    }
    # Aggregations
    total_students = Student.objects.count()
    avg_students_per_class = Student.objects.values('class_name').annotate(count=Count('id')).aggregate(Avg('count'))
    
    context = {
        'students': students,
        'total_students': total_students,
        'search_query': search_query,
    }
    return render(request, 'students/student_list.html', context)

@login_required
def student_create(request):
    # CREATE operation
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    else:
        form = StudentForm()
    return render(request, 'students/student_form.html', {'form': form, 'title': 'Add Student'})

@login_required
def student_update(request, pk):
    # UPDATE operation
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'students/student_form.html', {'form': form, 'title': 'Edit Student'})

@login_required
def student_delete(request, pk):
    # DELETE operation
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        return redirect('student_list')
    return render(request, 'students/student_confirm_delete.html', {'student': student})

@login_required
def student_result(request, pk):
    # READ operation with related objects
    student = get_object_or_404(Student, pk=pk)
    results = student.results.all()
    
    # Complex QuerySet - get subjects not added yet
    existing_subject_ids = results.values_list('subject_id', flat=True)
    available_subjects = Subject.objects.exclude(id__in=existing_subject_ids)
    
    # Calculate total percentage and grade
    total_percentage = 0
    if results:
        total_percentage = sum(r.percentage() for r in results) / results.count()
    
    if request.method == 'POST':
        form = ResultForm(request.POST)
        if form.is_valid():
            result = form.save(commit=False)
            result.student = student
            result.save()
            return redirect('student_result', pk=pk)
    else:
        form = ResultForm()
    
    context = {
        'student': student,
        'results': results,
        'form': form,
        'available_subjects': available_subjects,
        'total_percentage': total_percentage,
    }
    return render(request, 'students/student_result.html', context)

@login_required
def dashboard(request):
    # Advanced QuerySets
    top_students = Student.objects.annotate(
        avg_marks=Avg('results__marks_obtained')
    ).filter(avg_marks__gte=80).order_by('-avg_marks')[:5]
    
    class_statistics = Student.objects.values('class_name').annotate(
        total_students=Count('id'),
        average_marks=Avg('results__marks_obtained'),
        highest_marks=Max('results__marks_obtained')
    )
    
    subjects_data = Subject.objects.annotate(
        avg_marks=Avg('result__marks_obtained'),
        pass_count=Count('result', filter=Q(result__marks_obtained__gte=40))
    )
    
    context = {
        'top_students': top_students,
        'class_statistics': class_statistics,
        'subjects_data': subjects_data,
    }
    return render(request, 'students/dashboard.html', context)

# Login view
def login_view(request):
    if request.user.is_authenticated:
        return redirect('student_list')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('student_list')
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'students/login.html')

# Logout view
def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def export_students_excel(request):
    """Export students data to Excel"""
    students = Student.objects.all().order_by('class_name', 'name')
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students List"
    
    # Headers
    headers = ['ID', 'Name', 'Roll No', 'Email', 'Class', 'Created Date']
    ws.append(headers)
    
    # Data
    for student in students:
        ws.append([
            student.id,
            student.name,
            student.roll_no,
            student.email,
            student.class_name,
            student.created_at.strftime('%Y-%m-%d %H:%M')
        ])
    
    # Column width adjust
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students_data.xlsx'
    wb.save(response)
    return response

@login_required
def export_results_excel(request):
    """Export all results to Excel"""
    results = Result.objects.select_related('student', 'subject').all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Headers
    headers = ['Student Name', 'Roll No', 'Class', 'Subject', 'Marks Obtained', 'Total Marks', 'Percentage', 'Grade']
    ws.append(headers)
    
    # Data
    for result in results:
        ws.append([
            result.student.name,
            result.student.roll_no,
            result.student.class_name,
            result.subject.name,
            result.marks_obtained,
            result.total_marks,
            f"{result.percentage():.2f}%",
            result.grade()
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=all_results.xlsx'
    wb.save(response)
    return response

@login_required
def export_class_report(request, class_name):
    """Export specific class results"""
    students = Student.objects.filter(class_name=class_name)
    results = Result.objects.filter(student__in=students).select_related('student', 'subject')
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{class_name} Report"
    
    headers = ['Student Name', 'Roll No', 'Subject', 'Marks', 'Percentage', 'Grade']
    ws.append(headers)
    
    for result in results:
        ws.append([
            result.student.name,
            result.student.roll_no,
            result.subject.name,
            result.marks_obtained,
            f"{result.percentage():.2f}%",
            result.grade()
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={class_name}_report.xlsx'
    wb.save(response)
    return response